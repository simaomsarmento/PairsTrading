import pandas as pd
import numpy as np
from openpyxl import load_workbook

# just set the seed for the random number generator
np.random.seed(107)

class DataProcessor:
    """
    This class contains a set of pairs trading strategies along
    with some auxiliary functions

    """

    def read_ticker_excel(self, path=None):
        """
        Assumes the relevant tickers are saved in an excel file.

        :param path: path to excel
        :return: df with tickers data, list with tickers
        """

        df = pd.read_excel(path)

        # remove duplicated
        unique_df = df[~df.duplicated(subset=['Ticker'], keep='first')].sort_values(['Ticker'])
        tickers = unique_df.Ticker.unique()

        return df, unique_df, tickers

    def dict_to_df(self, dataset, threshold=None):
        """
        Transforms a dictionary into a Dataframe

        :param dataset: dictionary containing tickers as keys and corresponding price series
        :param threshold: threshold for number of Nan Values
        :return: df with tickers as columns
        :return: df_clean with tickers as columns, and columns with null values dropped
        """

        first_count = True
        for k in dataset.keys():
            if dataset[k] is not None:
                if first_count:
                    df = dataset[k]
                    first_count = False
                else:
                    df = pd.concat([df, dataset[k]], axis=1)

        if threshold is not None:
            df_clean = self.remove_tickers_with_nan(df, threshold)
        else:
            df_clean = df

        return df, df_clean

    def remove_tickers_with_nan(self, df, threshold):
        """
        Removes columns with more than threshold null values
        """
        null_values = df.isnull().sum()
        null_values = null_values[null_values > 0]

        to_remove = list(null_values[null_values > threshold].index)
        df = df.drop(columns=to_remove)

        return df

    def get_return_series(self, df_prices):
        """
        This function calculates the return series of a given price series

        :param prices: time series with prices
        :return: return series
        """
        df_returns = df_prices.pct_change()
        df_returns = df_returns.iloc[1:]

        return df_returns

    def split_data(self, df_prices, training_dates, testing_dates, remove_nan=True):
        """
        This function splits a dataframe into training and validation sets
        :param df_prices: dataframe containing prices for all dates
        :param training_dates: tuple (training initial date, training final date)
        :param testing_dates: tuple (testing initial date, testing final date)
        :param remove_nan: flag to detect if nan values are to be removed

        :return: df with training prices
        :return: df with testing prices
        """
        if remove_nan:
            dataset_mask = ((df_prices.index >= training_dates[0]) &\
                            (df_prices.index <= testing_dates[1]))
            df_prices_dataset = df_prices[dataset_mask]
            print('Total of {} tickers'.format(df_prices_dataset.shape[1]))
            df_prices_dataset_without_nan = self.remove_tickers_with_nan(df_prices_dataset, 0)
            print('Total of {} tickers after removing tickers with Nan values'.format(
                df_prices_dataset_without_nan.shape[1]))
            df_prices = df_prices_dataset_without_nan.copy()

        train_mask = (df_prices.index <= training_dates[1])
        test_mask = (df_prices.index >= testing_dates[0])
        df_prices_train = df_prices[train_mask]
        df_prices_test = df_prices[test_mask]

        return df_prices_train, df_prices_test

    def append_df_to_excel(self, filename, df, sheet_name='Sheet1', startrow=None,
                           truncate_sheet=False,
                           **to_excel_kwargs):
        """
        Source: https://stackoverflow.com/questions/20219254/how-to-write-to-an-existing-excel-file-without-overwriting
        -data-using-pandas/47740262#47740262

        Append a DataFrame [df] to existing Excel file [filename]
        into [sheet_name] Sheet.
        If [filename] doesn't exist, then this function will create it.

        Parameters:
          filename : File path or existing ExcelWriter
                     (Example: '/path/to/file.xlsx')
          df : dataframe to save to workbook
          sheet_name : Name of sheet which will contain DataFrame.
                       (default: 'Sheet1')
          startrow : upper left cell row to dump data frame.
                     Per default (startrow=None) calculate the last row
                     in the existing DF and write to the next row...
          truncate_sheet : truncate (remove and recreate) [sheet_name]
                           before writing DataFrame to Excel file
          to_excel_kwargs : arguments which will be passed to `DataFrame.to_excel()`
                            [can be dictionary]

        Returns: None
        """

        # ignore [engine] parameter if it was passed
        if 'engine' in to_excel_kwargs:
            to_excel_kwargs.pop('engine')

        writer = pd.ExcelWriter(filename, engine='openpyxl')

        # Python 2.x: define [FileNotFoundError] exception if it doesn't exist
        #try:
        #    FileNotFoundError
        #except NameError:
        #    FileNotFoundError = IOError

        try:
            # try to open an existing workbook
            writer.book = load_workbook(filename)

            # get the last row in the existing Excel sheet
            # if it was not specified explicitly
            if startrow is None and sheet_name in writer.book.sheetnames:
                startrow = writer.book[sheet_name].max_row

            # truncate sheet
            if truncate_sheet and sheet_name in writer.book.sheetnames:
                # index of [sheet_name] sheet
                idx = writer.book.sheetnames.index(sheet_name)
                # remove [sheet_name]
                writer.book.remove(writer.book.worksheets[idx])
                # create an empty sheet [sheet_name] using old index
                writer.book.create_sheet(sheet_name, idx)

            # copy existing sheets
            writer.sheets = {ws.title: ws for ws in writer.book.worksheets}
        except FileNotFoundError:
            # file does not exist yet, we will create it
            pass

        if startrow is None:
            startrow = 0

        # write out the new sheet
        df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)

        # save the workbook
        writer.save()
